#!/usr/bin/env python3
'''btd-library-duplicates.py searches for potentially duplicate CARTS in a Rivendell Library.
It writes the possible duplicates to standard output unless directed otherwise with
the command line option '--output <FILE>'. It can also write an Excel-compatible
spreadsheet file using the '--excel' option.

This is a complete hack! You need to set the Rivendell database parameters below
(starting at the line with "cnx = mysql.connector.connect(").
'''

import sys
import re
import argparse
import csv
import mysql.connector
import openpyxl

def my_print(*p_args, **p_kwargs):
    '''my_print sends its outptuto STDERR.'''
    print(*p_args, **p_kwargs, file=sys.stderr)

def main():
    '''Identify potentially duplicate Carts in a Rivendell library.
    Output to stdout, a file, or an Excel spreadsheet.'''

    output_file = sys.stdout
    csvfile = None

    parser = argparse.ArgumentParser(
        prog="btd-library-duplicates",
        description='Identify potential duplicate tracks in the Rivendell Library.')

    parser.add_argument('-c', '--csv',
                        help='Generate a so-called comma-separated-value text file containing the possible duplicates.',
                        action='store_true')
    parser.add_argument('-d', '--database',
                        help='Specify the database name (default: Rivendell)',
                        default='Rivendell',
                        action='store')
    parser.add_argument('-e', '--excel',
                        help='Generate an Excel workbook containing the possible duplicates listing.',
                        action='store_true')
    parser.add_argument('-g', '--groups',
                        help='Specify one or more groups (separated by commas) in which to search for duplicates. \
                        Remember to use quotes if you include spaces between group names.',
                        action='store')
    parser.add_argument('-n', '--hostname',
                        help='Specify the database host name or IP address (default: localhost).',
                        default='localhost',
                        action='store')
    parser.add_argument('-o', '--output',
                        help='Name the output file for the list of possible duplicates.',
                        action='store')
    parser.add_argument('-p', '--password',
                        help='Specify the database user password (no default).',
                        required=True,
                        action='store')
    parser.add_argument('-u', '--user',
                        help='Specify the database username (default: rduser).',
                        default='rduser',
                        action='store')
    parser.add_argument('-V', '--version',
                        help='Display this app version string.',
                        action='version',
                        version='%(prog)s: 0.1.1')
    parser.add_argument('-v', '--verbose',
                        help='Be chatty about progress. Use up to 3 times for more chattiness.',
                        default=0,
                        action='count')

    args = parser.parse_args()

    verbose_print = my_print if args.verbose > 0 else lambda *a, **k: None
    very_verbose_print = my_print if args.verbose > 1 else lambda *a, **k: None

    database_host = None
    database_name = None
    database_user = None
    database_passwd = None

    if args.hostname:
        database_host = args.hostname
    if args.database:
        database_name = args.database
    if args.user:
        database_user = args.user
    if args.password:
        database_passwd = args.password
    verbose_print("database params: host: {h} user: {u} pass: {p} db: {d}".format(
        h=database_host, u=database_user, p=database_passwd, d=database_name))

    if args.excel and not args.output:
        parser.print_help()
        sys.exit(1)

    if args.csv and not args.output:
        csvfile = sys.stdout

    if args.csv and args.output:
        output_file = args.output

    if args.excel and args.output:
        output_file = args.output

    if args.output:
        pat = re.compile(r'(?P<path>.*/)*?(?P<base>[-_\w]+)(\.(?P<ext>(xlsx?|csv)))?', flags=re.IGNORECASE)
        if output_file is not None and output_file != sys.stdout:
            matches = pat.search(output_file)
            parts = matches.groupdict()
            output_file_pathname = parts['path']
            output_file_basename = parts['base']
            output_file_extension = parts['ext']
            verbose_print('pathname: {p} basename: {b} extension: {e}'.format(
                p=output_file_pathname, b=output_file_basename, e=output_file_extension))

    header = {
        'dup_num': 'Duplicate #',
        'dup_of': 'Duplicate Of',
        'dup_artist': 'Artist',
        'dup_album': 'Album',
        'dup_title': 'Title',
        'org_length': 'Orig Length',
        'dup_length': 'Dupl Length',
    }
    # Support creation of both Excel AND CSV at the same time.
    if args.excel:
        workbook = openpyxl.Workbook()
        workbook.active.title = 'DuplicateCarts'
        workbook_filename = '{path}{base}.xslx'.format(
            path=(output_file_pathname if output_file_pathname is not None else ''),
            base=output_file_basename)

    if args.csv and args.output:
        csv_filename = '{path}{base}.csv'.format(
            path=(output_file_pathname if output_file_pathname is not None else ''),
            base=output_file_basename)
        try:
            csvfile = open(csv_filename, 'w', newline='', encoding="utf-8")
        except IOError as e:
            print("Unable to open '{f}' for writing ({e}).".format(f=csv_filename, e=e),
                  file=sys.stderr)
    elif output_file == sys.stdout:
        csv_filename = output_file

    if args.csv:
        csv_fields = ['dup_num',
                      'dup_of',
                      'dup_artist',
                      'dup_album',
                      'dup_title',
                      'org_length',
                      'dup_length',]
        file_writer = csv.DictWriter(
            csvfile,
            fieldnames=csv_fields,
            extrasaction='raise',
            delimiter='|',
            quotechar='^',
            quoting=csv.QUOTE_MINIMAL)

    try:
        cnx = mysql.connector.connect(
            host=database_host,
            user=database_user,
            passwd=database_passwd,
            database=database_name)
    except mysql.connector.Error as err:
        print("Error connecting to the database. Did you remember to edit this script? ({e})".format(e=err))

    cursor = cnx.cursor(dictionary=True)

    groups_list = []

    query = "select c.NUMBER as number, "
    query += "c.ARTIST as artist, c.ALBUM as album, c.TITLE as title, "
    query += "sec_to_time(u.LENGTH div 1000) as length "
    query += "from CART c join CUTS u on c.NUMBER = u.CART_NUMBER "
    # Audio carts only.
    query += "where c.TYPE = 1 "
    if args.groups is not None:
        groups_list = re.split(r', *', args.groups)
        query += "and c.GROUP_NAME in ({}) ".format(",".join(['%s'] * len(groups_list)))

    query += "order by c.ARTIST, c.ALBUM, c.TITLE"
    verbose_print("query: {q}".format(q=query))

    cursor.execute(query, groups_list)

    last_number = 0
    last_artist = ""
    last_album = ""
    last_title = ""
    last_length = 0

    # Output the Header Row.
    if args.excel:
        workbook.active.cell(row=1, column=1, value="Duplicate #")
        workbook.active.cell(row=1, column=2, value="Duplicate of")
        workbook.active.cell(row=1, column=3, value="Artist")
        workbook.active.cell(row=1, column=4, value="Album")
        workbook.active.cell(row=1, column=5, value="Track Title")
        workbook.active.cell(row=1, column=6, value='Orig Length')
        workbook.active.cell(row=1, column=7, value="Dupl Length")
    if args.csv:
        file_writer.writerow(header)
    if not args.output:
        print("{}|{}|{}|{}|{}|{}|{}".format(
            "Duplicate #",
            "Duplicate of",
            "Artist",
            "Album",
            "Track Title",
            "Orig Length",
            "Dupl Length"))

    row = 2
    d_row = {}
    for d_row in cursor:
        # This is the test for duplicity:
        if (
                d_row['number'] != last_number and
                d_row['title'] is not None and
                d_row['artist'] is not None and
                d_row['album'] == last_album and
                d_row['artist'] == last_artist and
                d_row['title'] == last_title
        ):
            row_data = {
                'dup_num':    d_row['number'],
                'dup_of':     last_number,
                'dup_artist': d_row['artist'],
                'dup_album':  d_row['album'],
                'dup_title':  d_row['title'],
                'org_length': last_length,
                'dup_length': d_row['length'].seconds + \
                d_row['length'].microseconds/1000000 if d_row['length'] is not None else 0,
            }
            very_verbose_print("row:{r}, data:{d}".format(r=row, d=row_data))

            if args.excel:
                workbook.active.cell(row=row, column=1, value=row_data['dup_num'])
                workbook.active.cell(row=row, column=2, value=last_number)
                workbook.active.cell(row=row, column=3, value=row_data['dup_artist'])
                workbook.active.cell(row=row, column=4, value=row_data['dup_album'])
                workbook.active.cell(row=row, column=5, value=row_data['dup_title'])
                workbook.active.cell(row=row, column=6, value=row_data['org_length'])
                workbook.active.cell(row=row, column=7, value=row_data['dup_length'])

            if args.csv:
                try:
                    file_writer.writerow(row_data)
                except ValueError as e:
                    print("Error writing track '{title}' to CSV file: {e}".format(title=row_data['dup_title'], e=e),
                          file=sys.stderr)

            row += 1

        last_number = d_row['number']
        last_artist = d_row['artist']
        last_album = d_row['album']
        last_title = d_row['title']
        last_length = d_row['length'].seconds + \
            d_row['length'].microseconds/1000000 if d_row['length'] is not None else 0

    if args.excel:
        try:
            workbook.save(filename=workbook_filename)
        except IOError as e:
            print("Error saving Excel file '{f}': {e}".format(f=workbook_filename, e=e),
                  file=sys.stderr)
    if args.csv:
        csvfile.close()

if __name__ == "__main__":
    main()
